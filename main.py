import requests
import json
from bs4 import BeautifulSoup
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side

with open('config.json', 'r', encoding='utf-8') as f:
    config = json.load(f)

def col_letter_to_index(col_letter):
    result = 0
    for char in col_letter:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result

def translate_age_to_category(age):
    match age:
        case _ if age <= 11:
            return "Dziecko"
        case _ if age > 11 and age <= 13:
            return "Mlodzik"
        case _ if age > 13 and age <= 16:
            return "Junior Mlodszy"
        case _ if age > 16 and age <= 18:
            return "Junior"
        case _ if age > 18 and age <= 23:
            return "Mlodzieżowiec"
        case _ if age > 23:
            return "Senior"
        
def translate_stroke(distance):
    distance_split = distance.split(" ")
    return f"{distance_split[0]} {config['styles'][distance_split[-1]]}"
        
def translate_date(date):
    date_list = date.split(" ")
    return f"{date_list[0]} {config['months'][date_list[1]]} {date_list[2]}"

clubId = config['clubId']
stroke = config['stroke']
courses = config['courses']
base_url = config['base_url']
ranking_details = "page=rankingDetail"
meetings = {}
regex_class = re.compile(r"^meetResult[0-1]$")
current_year = datetime.now().year
athlete_name = ""
athletes_dict = dict()

for course in courses:
    url = f"{base_url}?{ranking_details}&clubId={clubId}&stroke={stroke}&year={current_year}&course={course}"  
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    elements_td = soup.find_all('td', class_='name')

    for element_td in elements_td:
        link_element = element_td.find('a')
        if link_element:
            meeting_url = f"{base_url}{link_element.get('href')}&clubId={clubId}"
            meeting_response = requests.get(meeting_url)
            meeting_soup = BeautifulSoup(meeting_response.text, 'html.parser')
            meeting_name = meeting_soup.find('td', class_='titleLeft').decode_contents()
            meeting_city = meeting_soup.find_all('td', class_='titleLeft')[1].decode_contents().replace("\u00a0"," ")
            meeting_date = meeting_soup.find_all('td', class_='titleRight')[1].decode_contents().replace("\u00a0"," ")

            meetings[meeting_name] = {
                "city": meeting_city,
                "date": translate_date(meeting_date)
            }

            meeting_elements = meeting_soup.find_all('tr', class_=regex_class)
            for meeting_element in meeting_elements:
                meeting_class = meeting_element.get('class')[0]
                if meeting_class == "meetResult1":
                    athlete_element = meeting_element.find('td', class_='nameImportant')
                    if athlete_element:
                        athlete_name = str(athlete_element.find('a').decode_contents()).replace(",", "").title()
                        athlete_birth = translate_age_to_category(int(current_year) - int(athlete_element.decode_contents().split(" - ")[-1].strip()))
                        meetings[meeting_name][athlete_name] = {
                            "age": athlete_birth,
                            "numberOfStarts": 0
                        }
                        if athlete_name not in athletes_dict:
                            athletes_dict[athlete_name] = 0
                elif meeting_class == "meetResult0":
                    if meeting_element.find('td', class_='name').find('a') is not None:
                        distance_element = meeting_element.find('td', class_='name').find('a').decode_contents()
                        place_element = meeting_element.find('td', class_='meetPlace').decode_contents().split(" ")[-1]
                        if "Split" in place_element or "DSQ" in place_element or "4 x " in distance_element:
                            continue
                        place = place_element.replace(".", "")
                        if "swimmingEvents" not in meetings[meeting_name][athlete_name]:
                            meetings[meeting_name][athlete_name]["swimmingEvents"] = {}
                        meetings[meeting_name][athlete_name]["swimmingEvents"][translate_stroke(distance_element)] = place
                        meetings[meeting_name][athlete_name]["numberOfStarts"] += 1

most_events_per_athlete = {}
for meeting_name, data in meetings.items():
    for key in data.keys():
        if key in ["city", "date"]:
            continue
        athlete = key
        num_starts = data[athlete]["numberOfStarts"]
        age = data[athlete]["age"]

        if (athlete not in most_events_per_athlete or 
            num_starts > most_events_per_athlete[athlete]["numberOfStarts"]):
            most_events_per_athlete[athlete] = {
                "numberOfStarts": num_starts,
                "age": age
            }

wb = Workbook()
ws = wb.active
last_col_index = (len(meetings) + 1) * 3
last_col_letter = get_column_letter(last_col_index)
for i in range(1,4):
    ws.merge_cells(f"A{i}:{last_col_letter}{i}")
ws["A1"] = f"UCZESTNICTWO W ZAWODACH OBJĘTYCH SYSTEMEM WSPÓŁZAWODNICTWA SPORTOWEGO W {current_year} ROKU"
ws["A2"] = f"Nazwa klubu: Ks Posnania Poznan"
ws["A3"] = f"Dyscyplina: PŁYWANIE"
for i in range(1,5):
    ws.merge_cells(f"{get_column_letter(i)}4:{get_column_letter(i)}7")
ws["A4"] = "L.p. (*)"
ws["B4"] = f"Imię i nazwisko zawodnika objętego szkoleniem w {current_year} roku"
ws["C4"] = "Kategoria wiekowa"
ws["D4"] = "Liczba startów"

col_idx = 5
for meeting_name, meeting_info in meetings.items():
    col_letter_start = get_column_letter(col_idx)
    col_letter_end = get_column_letter(col_idx + 2)

    ws.merge_cells(f"{col_letter_start}4:{col_letter_end}4")
    ws[f"{col_letter_start}4"] = meeting_name
    ws.merge_cells(f"{col_letter_start}5:{col_letter_end}5")
    ws[f"{col_letter_start}5"] = meeting_info['date']
    ws.merge_cells(f"{col_letter_start}6:{col_letter_end}6")
    ws[f"{col_letter_start}6"] = meeting_info['city']

    ws[f"{col_letter_start}7"] = "Udział (*)"
    ws[f"{get_column_letter(col_idx + 1)}7"] = "Konkurencja"
    ws[f"{get_column_letter(col_idx + 2)}7"] = "Zajęte miejsce"

    col_idx += 3

sorted_athletes = sorted(most_events_per_athlete.items(), key=lambda x: x[0].split()[-1])
athlete_row_map = {}
row = 8

for i, (athlete, data) in enumerate(sorted_athletes, start=1):
    athlete_row_map[athlete] = row
    ws[f"A{row}"] = i
    ws[f"B{row}"] = athlete
    ws[f"C{row}"] = data["age"]
    ws[f"D{row}"] = data["numberOfStarts"]
    row += data["numberOfStarts"]

for meeting_index, (meeting_name, data) in enumerate(meetings.items()):
    meeting_col = 5 + (meeting_index * 3)
    for athlete_name in data:
        if athlete_name in ["city", "date"]:
            continue
        if athlete_name not in athlete_row_map:
            continue

        row_start = athlete_row_map[athlete_name]
        events = data[athlete_name].get("swimmingEvents", {})
        if not events:
            continue

        r = row_start
        for event_name, place in events.items():
            ws[f"{get_column_letter(meeting_col+1)}{r}"] = event_name
            ws[f"{get_column_letter(meeting_col+2)}{r}"] = place
            ws[f"{get_column_letter(meeting_col)}{r}"] = 1
            r += 1

center_alignment = Alignment(horizontal="center", vertical="center")
medium_border = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

for col_cells in ws.columns:
    max_length = 0
    column_letter = get_column_letter(col_cells[0].column)  
    
    for cell in col_cells:
        if cell.value:
            cell.alignment = center_alignment
        if cell.row >= 4:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
    
    adjusted_width = max_length + 2
    ws.column_dimensions[column_letter].width = adjusted_width
    
last_col_index = col_letter_to_index(last_col_letter)

for r in range(1, row + 1):
    for col in range(1, last_col_index + 2):
        cell = ws.cell(row=r, column=col)
        cell.border = medium_border
wb.save("uczestnictwo_2025.xlsx")
print("✅ Plik Excel zapisany jako 'uczestnictwo_2025.xlsx'")
